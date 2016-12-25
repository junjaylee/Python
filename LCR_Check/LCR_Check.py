################################################
# V1 : Initialize Version
################################################
import sys
import ConfigParser
import logging
import ast
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors, PatternFill, fills

def getCellWithMergeLookup(varSheet, varCol, varRow):
	idx = '{0}{1}'.format(get_column_letter(varCol), varRow)
	for range_ in varSheet.merged_cell_ranges:
		cells = list(openpyxl.utils.rows_from_range(range_))
		for row in cells:
			if idx in row:
				return varSheet.cell(cells[0][0])
	return varSheet.cell(row = varRow, column = varCol)
def Func_ReadCfgData(varCfg_File, varSection, varKeyword):
	Data = varCfg_File.get(varSection,varKeyword)
	List = Data.split(',')
	List[0] = ast.literal_eval(List[0])
	List[1] = ast.literal_eval(List[1])
	return List 
def Func_CompareText(varSheet, varY, varX, varText,varLog):
	Cell = varSheet.cell(row=varY, column=varX)
	if (Cell.value == None or Cell.value != varText):
		ErrMsg = '{0} != {1} in {2} sheet'.format(Cell.value,varText,varSheet)
		print ErrMsg
		logging.error(ErrMsg)
		return True
	return False
def Func_CheckDuplicatedKey(varSrcSheet, varDestSheet, varX1, varY1, varX2, varY2, varMarkLoc):
	Src_Series = pd.Series()
	Dest_Series = pd.Series()
	for iRow1 in range(varY1+1,varSrcSheet.max_row+1): 
		Cell1 = getCellWithMergeLookup(varSrcSheet,varX1,iRow1)
		if (Cell1.value == None):
			continue
		if (varMarkLoc != None):
			Cell = varSrcSheet.cell(row=iRow1, column=varMarkLoc)
			Cell.value = 'N'
		szValue1 = '%s' % (Cell1.value)
		for iRow2 in range(varY2+1,varDestSheet.max_row+1):
			Cell2 = getCellWithMergeLookup(varDestSheet,varX2,iRow2)
			if (Cell2.value == None):
				continue
			szValue2 = '%s' % (Cell2.value)
			if (szValue2.find(szValue1) != -1):
				if (varMarkLoc == None):
					Cell1 = varSrcSheet.cell(row=iRow1, column=varX1)
					Cell1.fill = PatternFill(fills.FILL_SOLID, fgColor=colors.RED)
					Cell2 = varDestSheet.cell(row=iRow2, column=varX2)
					Cell2.fill = PatternFill(fills.FILL_SOLID, fgColor=colors.RED)
				else:
					Cell.value = 'Y'
				Src_Series = Src_Series.append(pd.Series([szValue1]))
				Dest_Series = Dest_Series.append(pd.Series([szValue2]))
	return Src_Series, Dest_Series
def Func_CheckMissingKey(varSrcSheet, varDestSheet, varType1, varType2, varRefDes1, varRefDes2,varList,
	varDesc1, varDesc2):
	Rep_Data = pd.Series()
	Loc_Data = pd.Series()
	for iRow1 in range(varType1[1],varSrcSheet.max_row+1):
		Cell1 = getCellWithMergeLookup(varSrcSheet,varType1[0],iRow1)
		szValue1_Type = '%s' % (Cell1.value)
		if (Cell1.value == None or not (szValue1_Type in varList)):
			continue
		Cell2 = getCellWithMergeLookup(varSrcSheet,varRefDes1[0],iRow1)
		szValue1_RefDes = '%s' % (Cell2.value)
		if (Cell2.value == None):
			continue
		if (varDesc1 != None):
			Cell3 = getCellWithMergeLookup(varSrcSheet,varDesc1[0],iRow1)
			szKey1 = '%s' % (Cell3.value)
		words = szValue1_RefDes.split(',')
		for szKey in words:
			bFind = False
			for iRow2 in range(varType2[1],varDestSheet.max_row+1):
				Cell4 = getCellWithMergeLookup(varDestSheet,varType2[0],iRow2)
				szValue_Type = '%s' % (Cell4.value)
				if (Cell4 == None or not (szValue_Type in varList)):
					continue
				Cell5 = getCellWithMergeLookup(varDestSheet,varRefDes2[0],iRow2)
				szValue_RefDes = '%s' % (Cell5.value)
				if (Cell5.value == None):
					continue
				if (szValue_RefDes.find(szKey) != -1):
					if (varDesc2 != None):
						Cell6 = getCellWithMergeLookup(varDestSheet,varDesc2[0], iRow2)
						szKey2 = '%s' % (Cell6.value)
						if (szKey1 != szKey2):
							Loc_Data = Loc_Data.append(pd.Series([szKey]))
					bFind = True
					break
			if (bFind == False):
				Rep_Data = Rep_Data.append(pd.Series([szKey]))
	return Rep_Data, Loc_Data
#-----------------------------------------------
Tool_Version = 1.5
logging.basicConfig(filename='LCR_Check.log',level=logging.DEBUG)
logging.info('Version = %.2f', Tool_Version)
arg_total = len(sys.argv)
if (arg_total == 4):
	LCR_Check_File = sys.argv[1]
	BOM_File = sys.argv[2]
	BuildMatrix_File = sys.argv[3]
	logging.info('Input Paramters : Matrix => %s,\n LCR => %s,\n BOM => %s',BuildMatrix_File,LCR_Check_File,BOM_File)
else:
	LCR_Check_File = raw_input('Please input LCR measurement Report (Full Path)===>')
	BOM_File = raw_input('Please input TOP side and BOTTOM side BOM Report (Full Path)===>')
	BuildMatrix_File = raw_input('Please input Build Matrix File Name (Full Path) ===>')
	logging.info('Keyin Data : Matrix => %s, LCR => %s, BOM => %s',BuildMatrix_File,LCR_Check_File,BOM_File)
try:
	wb_Matrix = load_workbook(filename = BuildMatrix_File)
except IOError as e:
	logging.error('Open Matrix file (%d, %s)',e.errno , e.strerror)
	sys.exit(0)
try:
	wb_LCR = load_workbook(filename = LCR_Check_File)
except IOError as e:
	logging.error('Open LCR measurement file (%d, %s)',e.errno , e.strerror)
	sys.exit(0)
try:
	wb_BOM = load_workbook(filename = BOM_File)
except IOError as e:
	logging.error('Open BOM file (%d, %s)',e.errno , e.strerror)
	sys.exit(0)
#------------ Read Config.ini --------------------
config = ConfigParser.ConfigParser({'TOP_LCR':'TOP_LCR','BOT_LCR':'BOT_LCR','NOSTUFF_TOP':'NOSTUFF TOP','NOSTUFF_BOT':'NOSTUFF BOT'})
config.read('Config.ini')
sKey_TOP = config.get('LCR_WorkSheet', 'TOP_LCR')
sKey_BOT = config.get('LCR_WorkSheet', 'BOT_LCR')
sKey_NOSTUFF_TOP = config.get('LCR_WorkSheet', 'NOSTUFF_TOP')
sKey_NOSTUFF_BOT = config.get('LCR_WorkSheet', 'NOSTUFF_BOT')
iLoc_TOP_LCR = config.getint('LCR_WorkSheet', 'TOP_Desc_Loc')
iLoc_BOT_LCR = config.getint('LCR_WorkSheet', 'BOT_Desc_Loc')
iLoc_TOP_BOM = config.getint('BOM_WorkSheet', 'TOP_Desc_Loc')
iLoc_BOT_BOM = config.getint('BOM_WorkSheet', 'BOT_Desc_Loc')
iLoc_Matrix = config.getint('Matrix_WorkSheet', 'Desc_Loc')
ListTOP_Ref_Des = Func_ReadCfgData(config, 'LCR_WorkSheet', 'TOP_Ref_Des')
ListTOP_Item_Type = Func_ReadCfgData(config, 'LCR_WorkSheet', 'TOP_Item_Type')
ListTOP_Desc = Func_ReadCfgData(config, 'LCR_WorkSheet', 'TOP_Desc')
ListTOP_Matrix = Func_ReadCfgData(config, 'LCR_WorkSheet', 'TOP_Matrix')
ListBOT_Ref_Des = Func_ReadCfgData(config, 'LCR_WorkSheet', 'BOT_Ref_Des')
ListBOT_Item_Type = Func_ReadCfgData(config, 'LCR_WorkSheet', 'BOT_Item_Type')
ListBOT_Desc = Func_ReadCfgData(config, 'LCR_WorkSheet', 'BOT_Desc')
ListBOT_Matrix = Func_ReadCfgData(config, 'LCR_WorkSheet', 'BOT_Matrix')
ListNSTOP_Ref_Des = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_TOP_Ref_Des')
ListNSTOP_Item_Type = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_TOP_Item_Type')
ListNSTOP_Matrix = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_TOP_Matrix')
#ListNSTOP_Desc = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_TOP_Desc')
ListNSTOP_BOM = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_TOP_BOM')
ListNSBOT_Ref_Des = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_BOT_Ref_Des')
ListNSBOT_Item_Type = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_BOT_Item_Type')
ListNSBOT_Matrix = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_BOT_Matrix')
ListNSBOT_BOM = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_BOT_BOM')
#ListNSBOT_Desc = Func_ReadCfgData(config, 'LCR_WorkSheet', 'NS_BOT_Desc')
ListBOM_TOP_Ref_Des = Func_ReadCfgData(config, 'BOM_WorkSheet', 'TOP_Ref_Des')
ListBOM_TOP_Item_Type = Func_ReadCfgData(config, 'BOM_WorkSheet', 'TOP_Item_Type')
ListBOM_TOP_Desc = Func_ReadCfgData(config, 'BOM_WorkSheet', 'TOP_Desc')
ListBOM_BOT_Ref_Des = Func_ReadCfgData(config, 'BOM_WorkSheet', 'BOT_Ref_Des')
ListBOM_BOT_Item_Type = Func_ReadCfgData(config, 'BOM_WorkSheet', 'BOT_Item_Type')
ListBOM_BOT_Desc = Func_ReadCfgData(config, 'BOM_WorkSheet', 'BOT_Desc')
ListMatrix_Conf = Func_ReadCfgData(config, 'Matrix_WorkSheet', 'Sticky')
ListMatrix_Ref_Des = Func_ReadCfgData(config, 'Matrix_WorkSheet', 'Ref_Des')
ListMatrix_Desc = Func_ReadCfgData(config, 'Matrix_WorkSheet', 'Desc')
sKey_Type = config.get('LCR_WorkSheet', 'Type_List')
sKey_SMDT = config.get('BOM_WorkSheet', 'SMDT')
sKey_SMDB = config.get('BOM_WorkSheet', 'SMDB')
sKey_Matrix = config.get('Matrix_WorkSheet', 'WorkSheet')
words_Type = sKey_Type.split(',')
logging.info('sKey_TOP = %s, sKey_BOT = %s, sKey_NOSTUFF_TOP = %s, sKey_NOSTUFF_BOT = %s,words_Type = %s,sKey_SMDT = %s,sKey_SMDB = %s,sKey_Matrix = %s,', 
	sKey_TOP, sKey_BOT, sKey_NOSTUFF_TOP, sKey_NOSTUFF_BOT,words_Type,sKey_SMDT,sKey_SMDB,sKey_Matrix)
logging.info('ListTOP_Ref_Des = %s,ListTOP_Item_Type = %s,ListTOP_Desc = %s,ListBOT_Ref_Des = %s,ListBOT_Item_Type = %s,ListBOT_Desc = %s',
	ListTOP_Ref_Des,ListTOP_Item_Type,ListTOP_Desc,ListBOT_Ref_Des,ListBOT_Item_Type,ListBOT_Desc)
logging.info('ListNSTOP_Ref_Des = %s,ListNSBOT_Ref_Des = %s,ListMatrix_Conf = %s,ListMatrix_Ref_Des = %s,ListMatrix_Desc = %s',
	ListNSTOP_Ref_Des,ListNSBOT_Ref_Des,ListMatrix_Conf,ListMatrix_Ref_Des,ListMatrix_Desc)
logging.info('ListTOP_Matrix = %s,ListBOT_Matrix = %s,iLoc_TOP_LCR = %d,iLoc_BOT_LCR = %d,iLoc_TOP_BOM = %d,iLoc_BOT_BOM = %d,iLoc_Matrix = %d',
	ListTOP_Matrix,ListBOT_Matrix,iLoc_TOP_LCR,iLoc_BOT_LCR,iLoc_TOP_BOM,iLoc_BOT_BOM,iLoc_Matrix)
logging.info('ListBOM_TOP_Desc = %s,ListBOM_BOT_Desc = %s,ListNSTOP_Matrix = %s,ListNSBOT_Matrix = %s,ListNSTOP_BOM = %s,ListNSBOT_BOM = %s',
	ListBOM_TOP_Desc,ListBOM_BOT_Desc,ListNSTOP_Matrix,ListNSBOT_Matrix,ListNSTOP_BOM,ListNSBOT_BOM)
logging.info('ListNSTOP_Item_Type = %s,ListNSBOT_Item_Type = %s',ListNSTOP_Item_Type,ListNSBOT_Item_Type)
wsBOM_SMDT = wb_BOM[sKey_SMDT]
wsBOM_SMDB = wb_BOM[sKey_SMDB]
wsMatrix = wb_Matrix[sKey_Matrix]
logging.info('sKey_SMDT = %s, wsBOM_SMDT = %s, sKey_SMDB = %s, wsBOM_SMDB = %s,',sKey_SMDT,wsBOM_SMDT,sKey_SMDB,wsBOM_SMDB)
#---------- Check Duplicated Symbol ------------
wsTOP = wb_LCR[sKey_TOP]
wsBOT = wb_LCR[sKey_BOT]
wsNOSTUFF_TOP = wb_LCR[sKey_NOSTUFF_TOP]
wsNOSTUFF_BOT = wb_LCR[sKey_NOSTUFF_BOT]
#--------------- Check TOP Side ----------------
if (Func_CompareText(wsNOSTUFF_TOP, ListNSTOP_Ref_Des[1], ListNSTOP_Ref_Des[0], ListNSTOP_Ref_Des[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsTOP, ListTOP_Ref_Des[1], ListTOP_Ref_Des[0], ListTOP_Ref_Des[2], logging)):
	sys.exit(0)
Series_NOSTUFF, Series_TOP = Func_CheckDuplicatedKey(wsNOSTUFF_TOP,wsTOP,ListNSTOP_Ref_Des[0],
	ListNSTOP_Ref_Des[1],ListTOP_Ref_Des[0],ListTOP_Ref_Des[1],None)
TOP_DataFrame = pd.DataFrame({ 'NOSTUFF_TOP' : Series_NOSTUFF,'TOP' : Series_TOP})
print 'TOP LCR Compare Result =>\n{0}'.format(TOP_DataFrame)
logging.info('TOP LCR Compare Result =>\n%s',TOP_DataFrame)
#--------------- Check BOTTOM Side ----------------
if (Func_CompareText(wsNOSTUFF_BOT, ListNSBOT_Ref_Des[1], ListNSBOT_Ref_Des[0], ListNSBOT_Ref_Des[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsBOT, ListBOT_Ref_Des[1], ListBOT_Ref_Des[0], ListBOT_Ref_Des[2], logging)):
	sys.exit(0)
Series_NOSTUFF, Series_BOT = Func_CheckDuplicatedKey(wsNOSTUFF_BOT,wsBOT,ListNSBOT_Ref_Des[0],
	ListNSBOT_Ref_Des[1],ListBOT_Ref_Des[0],ListBOT_Ref_Des[1],None)
BOT_DataFrame = pd.DataFrame({ 'NOSTUFF_BOT' : Series_NOSTUFF,'BOT' : Series_BOT})
print 'BOT LCR Compare Result =>\n{0}'.format(BOT_DataFrame)
logging.info('BOT LCR Compare Result =>\n%s',BOT_DataFrame)
#------------------ Check BOM TOP -----------------
if (Func_CompareText(wsTOP, ListTOP_Item_Type[1], ListTOP_Item_Type[0], ListTOP_Item_Type[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsTOP, ListTOP_Desc[1], ListTOP_Desc[0], ListTOP_Desc[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsNOSTUFF_TOP, ListNSTOP_Item_Type[1], ListNSTOP_Item_Type[0], ListNSTOP_Item_Type[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsNOSTUFF_TOP, ListNSTOP_BOM[1], ListNSTOP_BOM[0], ListNSTOP_BOM[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsBOM_SMDT, ListBOM_TOP_Item_Type[1], ListBOM_TOP_Item_Type[0], ListBOM_TOP_Item_Type[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsBOM_SMDT, ListBOM_TOP_Ref_Des[1], ListBOM_TOP_Ref_Des[0], ListBOM_TOP_Ref_Des[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsBOM_SMDT, ListBOM_TOP_Desc[1], ListBOM_TOP_Desc[0], ListBOM_TOP_Desc[2], logging)):
	sys.exit(0)
Missing_BOM, DescFail_BOM = Func_CheckMissingKey(wsTOP,wsBOM_SMDT,[ListTOP_Item_Type[0],ListTOP_Item_Type[1]],[ListBOM_TOP_Item_Type[0],ListBOM_TOP_Item_Type[1]],
	[ListTOP_Ref_Des[0],ListTOP_Ref_Des[1]],[ListBOM_TOP_Ref_Des[0],ListBOM_TOP_Ref_Des[1]],words_Type,[ListTOP_Desc[0], ListTOP_Desc[1]],
	[ListBOM_TOP_Desc[0], ListBOM_TOP_Desc[1]])
print 'In LCR TOP But missing in TOP BOM =\n{0}'.format(Missing_BOM)
logging.info('In LCR TOP But missing in TOP BOM =\n%s',Missing_BOM)
print 'Descript can not match TOP LCR and BOM = \n{0}'.format(DescFail_BOM)
logging.info('In LCR TOP But missing in TOP BOM =\n%s',DescFail_BOM)
Missing_TOP, DescFail_BOM = Func_CheckMissingKey(wsBOM_SMDT,wsTOP,[ListBOM_TOP_Item_Type[0],ListBOM_TOP_Item_Type[1]],[ListTOP_Item_Type[0],ListTOP_Item_Type[1]],
	[ListBOM_TOP_Ref_Des[0],ListBOM_TOP_Ref_Des[1]],[ListTOP_Ref_Des[0],ListTOP_Ref_Des[1]],words_Type,None,None)
print 'In TOP BOM But missing in LCR TOP =\n{0}'.format(Missing_TOP)
logging.info('In TOP BOM But missing in LCR TOP =\n%s',Missing_TOP)
Series_NSTOP, Series_TOPBOM = Func_CheckDuplicatedKey(wsNOSTUFF_TOP,wsBOM_SMDT,ListNSTOP_Ref_Des[0],
	ListNSTOP_Ref_Des[1],ListBOM_TOP_Ref_Des[0],ListBOM_TOP_Ref_Des[1],ListNSTOP_BOM[0])
print 'In LCR NOSTUFF TOP But found in TOP BOM =\n{0}'.format(Series_NSTOP)
logging.info('In LCR NOSTUFF TOP But found in TOP BOM =\n%s',Series_NSTOP)
#------------------ Check BOM BOM -----------------
if (Func_CompareText(wsBOT, ListBOT_Item_Type[1], ListBOT_Item_Type[0], ListBOT_Item_Type[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsNOSTUFF_BOT, ListNSBOT_Item_Type[1], ListNSBOT_Item_Type[0], ListNSBOT_Item_Type[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsNOSTUFF_BOT, ListNSBOT_BOM[1], ListNSBOT_BOM[0], ListNSBOT_BOM[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsBOM_SMDB, ListBOM_BOT_Item_Type[1], ListBOM_BOT_Item_Type[0], ListBOM_BOT_Item_Type[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsBOM_SMDB, ListBOM_BOT_Ref_Des[1], ListBOM_BOT_Ref_Des[0], ListBOM_BOT_Ref_Des[2], logging)):
	sys.exit(0)
Missing_BOM, DescFail_BOM = Func_CheckMissingKey(wsBOT,wsBOM_SMDB,[ListBOT_Item_Type[0],ListBOT_Item_Type[1]],[ListBOM_BOT_Item_Type[0],ListBOM_BOT_Item_Type[1]],
	[ListBOT_Ref_Des[0],ListBOT_Ref_Des[1]],[ListBOM_BOT_Ref_Des[0],ListBOM_BOT_Ref_Des[1]],words_Type,[ListBOT_Desc[0], ListBOT_Desc[1]],
	[ListBOM_BOT_Desc[0],ListBOM_BOT_Desc[1]])
print 'In LCR BOT But missing in BOT BOM =\n{0}'.format(Missing_BOM)
logging.info('In LCR BOT But missing in BOT BOM =\n%s',Missing_BOM)
print 'Descript can not match BOT LCR and BOM = \n{0}'.format(DescFail_BOM)
logging.info('In LCR BOT But missing in BOT BOM =\n%s',DescFail_BOM)
Missing_BOT, DescFail_BOM = Func_CheckMissingKey(wsBOM_SMDB,wsBOT,[ListBOM_BOT_Item_Type[0],ListBOM_BOT_Item_Type[1]],[ListBOT_Item_Type[0],ListBOT_Item_Type[1]],
	[ListBOM_BOT_Ref_Des[0],ListBOM_BOT_Ref_Des[1]],[ListBOT_Ref_Des[0],ListBOT_Ref_Des[1]],words_Type,None,None)
print 'In BOT BOM But missing in LCR BOT =\n{0}'.format(Missing_BOT)
logging.info('In BOT BOM But missing in LCR BOT =\n%s',Missing_BOT)
Series_NSBOT, Series_BOTBOM = Func_CheckDuplicatedKey(wsNOSTUFF_BOT,wsBOM_SMDB,ListNSBOT_Ref_Des[0],
	ListNSBOT_Ref_Des[1],ListBOM_BOT_Ref_Des[0],ListBOM_BOT_Ref_Des[1],ListNSBOT_BOM[0])
print 'In LCR NOSTUFF BOT But found in BOT BOM =\n{0}'.format(Series_NSBOT)
logging.info('In LCR NOSTUFF BOT But found in BOT BOM =\n%s',Series_NSBOT)
#------------------ Check Matrix TOP -----------------
if (Func_CompareText(wsMatrix, ListMatrix_Conf[1], ListMatrix_Conf[0], ListMatrix_Conf[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsMatrix, ListMatrix_Ref_Des[1], ListMatrix_Ref_Des[0], ListMatrix_Ref_Des[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsMatrix, ListMatrix_Desc[1], ListMatrix_Desc[0], ListMatrix_Desc[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsTOP, ListTOP_Desc[1], ListTOP_Desc[0], ListTOP_Desc[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsBOT, ListBOT_Desc[1], ListBOT_Desc[0], ListBOT_Desc[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsTOP, ListTOP_Matrix[1], ListTOP_Matrix[0], ListTOP_Matrix[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsBOT, ListBOT_Matrix[1], ListBOT_Matrix[0], ListBOT_Matrix[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsNOSTUFF_TOP, ListNSTOP_Ref_Des[1], ListNSTOP_Ref_Des[0], ListNSTOP_Ref_Des[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsNOSTUFF_BOT, ListNSBOT_Ref_Des[1], ListNSBOT_Ref_Des[0], ListNSBOT_Ref_Des[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsNOSTUFF_TOP, ListNSTOP_Matrix[1], ListNSTOP_Matrix[0], ListNSTOP_Matrix[2], logging)):
	sys.exit(0)
if (Func_CompareText(wsNOSTUFF_BOT, ListNSBOT_Matrix[1], ListNSBOT_Matrix[0], ListNSBOT_Matrix[2], logging)):
	sys.exit(0)
Matrix_Conf_Series = pd.Series()
Matrix_Ref_Des_Series = pd.Series()
Matrix_Desc_Series = pd.Series()
for iRow in range(ListMatrix_Ref_Des[1],wsMatrix.max_row+1):
	Cell = getCellWithMergeLookup(wsMatrix,ListMatrix_Conf[0],iRow)
	if (Cell == None):
		Matrix_Conf_Series = Matrix_Conf_Series.append(pd.Series([""]))
	else:
		Matrix_Conf_Series = Matrix_Conf_Series.append(pd.Series([Cell.value]))
	Cell = getCellWithMergeLookup(wsMatrix,ListMatrix_Ref_Des[0],iRow)
	if (Cell == None):
		Matrix_Ref_Des_Series = Matrix_Ref_Des_Series.append(pd.Series([""]))
	else:
		Matrix_Ref_Des_Series = Matrix_Ref_Des_Series.append(pd.Series([Cell.value]))
	Cell = getCellWithMergeLookup(wsMatrix,ListMatrix_Desc[0],iRow)
	if (Cell == None):
		Matrix_Desc_Series = Matrix_Desc_Series.append(pd.Series([""]))
	else:
		Matrix_Desc_Series = Matrix_Desc_Series.append(pd.Series([Cell.value]))
print 'Load Matrix data to memory, It has {0}-{1}-{2} records'.format(len(Matrix_Conf_Series),len(Matrix_Ref_Des_Series),len(Matrix_Desc_Series))
TOP_Series = pd.Series()
BOT_Series = pd.Series()
Desc_Series = pd.Series()
NSTOP_Series = pd.Series()
NSBOT_Series = pd.Series()
Y_TOP_Ref_Des = ListTOP_Ref_Des[1]
Y_BOT_Ref_Des = ListBOT_Ref_Des[1]
Y_NSTOP_Ref_Des = ListNSTOP_Ref_Des[1]
Y_NSBOT_Ref_Des = ListNSBOT_Ref_Des[1]
Max_Row_TOP = wsTOP.max_row+1
Max_Row_BOT = wsBOT.max_row+1
Max_Row_NSTOP = wsNOSTUFF_TOP.max_row+1
Max_Row_NSBOT = wsNOSTUFF_BOT.max_row+1
while  (Y_TOP_Ref_Des <= Max_Row_TOP or Y_BOT_Ref_Des <= Max_Row_BOT or Y_NSBOT_Ref_Des <= Max_Row_NSBOT or Y_NSTOP_Ref_Des <= Max_Row_NSTOP):
	Y_TOP_Ref_Des += 1
	Y_BOT_Ref_Des += 1
	Y_NSTOP_Ref_Des += 1
	Y_NSBOT_Ref_Des += 1
	logging.info('Progess TOP : %d, BOT : %d',Y_TOP_Ref_Des,Y_BOT_Ref_Des)
	Cell = getCellWithMergeLookup(wsTOP,ListTOP_Ref_Des[0],Y_TOP_Ref_Des)
	if (Cell.value == None):
		szValue_TOP = None
	else:	
		szValue_TOP = '%s' % (Cell.value)
	Cell = getCellWithMergeLookup(wsTOP,ListTOP_Desc[0],Y_TOP_Ref_Des)
	if (Cell.value == None):
		szDesc_TOP = None
	else:	
		szDesc_TOP = '%s' % (Cell.value)
	Cell = wsTOP.cell(row=Y_TOP_Ref_Des, column=ListTOP_Matrix[0])
	Cell.value = ''
	Cell = getCellWithMergeLookup(wsBOT,ListBOT_Ref_Des[0],Y_BOT_Ref_Des)
	if (Cell.value == None):
		szValue_BOT = None
	else:
		szValue_BOT = '%s' % (Cell.value)
	Cell = getCellWithMergeLookup(wsBOT,ListBOT_Desc[0],Y_BOT_Ref_Des)
	if (Cell.value == None):
		szDesc_BOT = None
	else:
		szDesc_BOT = '%s' % (Cell.value)
	Cell = wsBOT.cell(row=Y_BOT_Ref_Des, column=ListBOT_Matrix[0])
	Cell.value = ''
	Cell = getCellWithMergeLookup(wsNOSTUFF_TOP,ListNSTOP_Ref_Des[0],Y_NSTOP_Ref_Des)
	if (Cell.value == None):
		szValue_NSTOP = None
	else:
		szValue_NSTOP = '%s' % (Cell.value)
	Cell = wsNOSTUFF_TOP.cell(row=Y_NSTOP_Ref_Des, column=ListNSTOP_Matrix[0])
	Cell.value = ''
	# Cell = getCellWithMergeLookup(wsNOSTUFF_TOP,ListNSTOP_Desc[0],Y_NSTOP_Ref_Des)
	# if (Cell.value == None):
	# 	szDesc_BOT = None
	# else:
	# 	szDesc_NSTOP = '%s' % (Cell.value)
	Cell = getCellWithMergeLookup(wsNOSTUFF_BOT,ListNSBOT_Ref_Des[0],Y_NSBOT_Ref_Des)
	if (Cell.value == None):
		szValue_NSBOT = None
	else:
		szValue_NSBOT = '%s' % (Cell.value)
	# Cell = getCellWithMergeLookup(wsNOSTUFF_BOT,ListNSBOT_Desc[0],Y_NSBOT_Ref_Des)
	# if (Cell.value == None):
	# 	szDesc_NSBOT = None
	# else:
	# 	szDesc_NSBOT = '%s' % (Cell.value)
	Cell = wsNOSTUFF_BOT.cell(row=Y_NSBOT_Ref_Des, column=ListNSBOT_Matrix[0])
	Cell.value = ''
	if (szValue_BOT == None and szValue_TOP == None and szValue_NSBOT == None and szValue_NSTOP == None):
		continue
	for iRow in range(0,len(Matrix_Conf_Series)-1):
		szValue_Conf = Matrix_Conf_Series.iloc[iRow]
		if (szValue_Conf != 1):
			continue
		szValue_RefDes = Matrix_Ref_Des_Series.iloc[iRow]
		szDesc_Matrix = Matrix_Desc_Series.iloc[iRow]
		bFind = False
		if (szValue_TOP != None):
			for sTOP in szValue_TOP.split(','):
				if (szValue_RefDes.find(sTOP) != -1 and sTOP != ''):
					bFind = True
					logging.info('LCR TOP [row=%d][%s] Control by Matrix [row=%d]',Y_TOP_Ref_Des,sTOP,iRow)
		if (bFind):
			Cell = wsTOP.cell(row=Y_TOP_Ref_Des, column=ListTOP_Matrix[0])
			if (Cell == None):
				szData = ''
			else:
				szData = '%s' % (Cell.value)
			if (szDesc_Matrix != szDesc_TOP):
				if (szData.find('Description not match') == -1):
					Cell.value = '{0}Description not match\n'.format(szData)
					szData = Cell.value
				logging.info('LCR TOP [row=%d][%s] Desc not match Matrix [row=%d][%s]',Y_TOP_Ref_Des,szDesc_TOP,iRow,szValue_TOP)
				Desc_Series = Desc_Series.append(pd.Series([szValue_TOP]))
			TOP_Series = TOP_Series.append(pd.Series([szValue_TOP]))	
			if (szData.find('Controlled by matrix') == -1):	
				Cell.value = '{0}Controlled by matrix'.format(szData)
		bFind = False
		if (szValue_BOT != None):
			for sBOT in szValue_BOT.split(','):
				if (szValue_RefDes.find(sBOT) != -1 and sBOT != ''):
					bFind = True
					logging.info('LCR BOT [row=%d][%s] Control by Matrix [row=%d]',Y_BOT_Ref_Des,sBOT,iRow)
		if (bFind):
			Cell = wsBOT.cell(row=Y_BOT_Ref_Des, column=ListBOT_Matrix[0])
			if (Cell == None):
				szData = ''
			else:
				szData = '%s' % (Cell.value)
			if (szDesc_Matrix != szDesc_BOT):
				if (szData.find('Description not match') == -1 or szData == ''):
					Cell.value = '{0}Description not match\n'.format(szData)
					szData = Cell.value
				logging.info('LCR BOT [row=%d][%s] Desc not match Matrix [row=%d][%s]',Y_BOT_Ref_Des,szDesc_BOT,iRow,szValue_BOT)
				Desc_Series = Desc_Series.append(pd.Series([szValue_BOT]))
			BOT_Series = BOT_Series.append(pd.Series([szValue_BOT]))
			if (szData.find('Controlled by matrix') == -1):
				Cell.value = '{0}Controlled by matrix'.format(szData)
		bFind = False
		if (szValue_NSTOP != None):
			for sNSTOP in szValue_NSTOP.split(','):
				if (szValue_RefDes.find(sNSTOP) != -1 and sNSTOP != ''):
					bFind = True
					logging.info('LCR NSTOP [row=%d][%s] Control by Matrix [row=%d]',Y_NSTOP_Ref_Des,sNSTOP,iRow)
		if (bFind):
			Cell = wsNOSTUFF_TOP.cell(row=Y_NSTOP_Ref_Des, column=ListNSTOP_Matrix[0])
			szData = '%s' % (Cell.value)
			# if (szDesc_Matrix != szDesc_NSTOP):
			# 	Cell.value = '{0}\nDesc not match'.format(szData)
			# 	Desc_Series = Desc_Series.append(pd.Series([szValue_NSTOP]))
			NSTOP_Series = NSTOP_Series.append(pd.Series([szValue_NSTOP]))
			Cell.value = 'Controlled by matrix'
		bFind = False
		if (szValue_NSBOT != None):
			for sNSBOT in szValue_NSBOT.split(','):
				if (szValue_RefDes.find(sNSBOT) != -1 and sNSBOT != ''):
					bFind = True
					logging.info('LCR NSBOT [row=%d][%s] Control by Matrix [row=%d]',Y_NSBOT_Ref_Des,sNSBOT,iRow)
		if (bFind):
			Cell = wsNOSTUFF_BOT.cell(row=Y_NSBOT_Ref_Des, column=ListNSBOT_Matrix[0])
			szData = '%s' % (Cell.value)
			# if (szDesc_Matrix != szDesc_NSBOT):
			# 	Cell.value = '{0}\nDesc not match'.format(szData)
			# 	Desc_Series = Desc_Series.append(pd.Series([szValue_NSBOT]))
			NSBOT_Series = NSBOT_Series.append(pd.Series([szValue_NSBOT]))
			Cell.value = 'Controlled by matrix'
TOP_Series = TOP_Series.drop_duplicates()
BOT_Series = BOT_Series.drop_duplicates()
NSTOP_Series = NSTOP_Series.drop_duplicates()
NSBOT_Series = NSBOT_Series.drop_duplicates()
Desc_Series = Desc_Series.drop_duplicates()
print 'LCR TOP Control by Matrix =\n{0}'.format(TOP_Series)
logging.info('LCR TOP Control by Matrix =\n%s',TOP_Series)
print 'LCR BOT Control by Matrix =\n{0}'.format(BOT_Series)
logging.info('LCR BOT Control by Matrix =\n%s',BOT_Series)
print 'LCR NSTOP Control by Matrix =\n{0}'.format(NSTOP_Series)
logging.info('LCR NSTOP Control by Matrix =\n%s',NSTOP_Series)
print 'LCR NSBOT Control by Matrix =\n{0}'.format(NSBOT_Series)
logging.info('LCR NSBOT Control by Matrix =\n%s',NSBOT_Series)
print 'Descript can not match LCR and Matrix = \n{0}'.format(Desc_Series)
logging.info('Descript can not match LCR and Matrix =\n%s',Desc_Series)
wb_LCR.save(filename = "LCR_Result.xlsx")
logging.info('Finished')
print 'Finished'